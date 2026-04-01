"""
Generate submission deliverables for both papers:
  1. Paper_Info_Record.pdf  (in each paper folder)
  2. paper_info.xlsx        (in full_run/ + repo root)
  3. regression_results.xlsx (in full_run/ + repo root)

Run from repo root:
  python generate_deliverables.py
"""

from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF

REPO_ROOT = Path(__file__).parent
WEEK2 = REPO_ROOT / "Week 2"   # rename to paper_analysis_output once OS allows

PAPERS = {
    "0005": {
        "paper_dir": WEEK2 / "Paper_0005_MappingEntrepreneurial",
        "workbook":  WEEK2 / "Paper_0005_MappingEntrepreneurial" / "full_run" / "Stroube2025Report_0005.xlsx",
        "pir_txt":   WEEK2 / "Paper_0005_MappingEntrepreneurial" / "Paper_Info_Record.txt",
    },
    "0017": {
        "paper_dir": WEEK2 / "Paper_0017_StatusConsensus",
        "workbook":  WEEK2 / "Paper_0017_StatusConsensus" / "full_run" / "Stroube2024Report_0017.xlsx",
        "pir_txt":   WEEK2 / "Paper_0017_StatusConsensus" / "Paper_Info_Record.txt",
    },
}

# ---------------------------------------------------------------------------
# Structured PIR data for both papers
# ---------------------------------------------------------------------------
PIR_DATA = {
    "0005": [
        ("Paper ID",               "0005"),
        ("Title",                  "Mapping Entrepreneurial Inclusion Across US Neighborhoods: The Case of Low-Code E-commerce Entrepreneurship"),
        ("Authors",                "Bryan Stroube, Gary Dushnitsky"),
        ("Journal",                "Strategic Management Journal (~2025)"),
        ("Data type",              "Cross-section (ZCTA-level)"),
        ("Observations (N)",       "32,647"),
        ("Preferred model column", "Table 3, Column 1 (Model 3_1)"),
        ("Dependent variable",     "log_shopify_count_1"),
        ("Focal IV",               "log_pop_black_aa"),
        ("Controls",               "log_pop_total, log_total_bachelor_deg, log_pop_total_poverty, log_total_social_cap"),
        ("Fixed effects",          "state_name_fe + MSA_fe (two-way, absorbed via pyfixest)"),
        ("Clustered SEs",          "CRV1 on MSA_fe"),
        ("Weights",                "None"),
        ("Baseline coefficient",   "0.0307 (log_pop_black_aa)"),
        ("Baseline SE",            "0.0054"),
        ("Baseline p-value",       "<0.001 (p = 2.80e-08)"),
        ("Published value",        "0.0307 (exact match)"),
        ("R-squared",              "0.688"),
        ("Key variables (4)",      "log_pop_black_aa, log_total_bachelor_deg, log_pop_total_poverty, log_total_social_cap"),
        ("MAR control",            "log_pop_total"),
        ("Mechanisms",             "MCAR, MAR, NMAR"),
        ("Missingness proportions","1%, 5%, 10%, 20%, 30%, 40%, 50%"),
        ("Methods",                "LD, Mean, Reg, Iter, RF, DL, MILGBM"),
        ("Iterations per scenario","30"),
        ("MAR/NMAR strength",      "1.5"),
        ("M (multiple imputation)","5"),
        ("DL method",              "TensorFlow 2.20.0, MLP 32->Dropout(0.1)->16->1, ReLU, Adam lr=0.005, EarlyStopping"),
        ("MILGBM method",          "LightGBM 4.6.0, MICE M=5, Rubin's Rules, 30 estimators, max_depth=4"),
        ("Full run start",         "2026-03-29 21:17"),
        ("Full run end",           "2026-04-01 07:38:40"),
        ("Total runs",             "17,640 (0 errors)"),
        ("QC status",              "PASSED (2026-04-01)"),
    ],
    "0017": [
        ("Paper ID",               "0017"),
        ("Title",                  "Status and Consensus: Heterogeneity in Audience Evaluations of Female- versus Male-Lead Films"),
        ("Authors",                "Bryan Stroube"),
        ("Journal",                "Strategic Management Journal, 45:994-1024 (2024)"),
        ("Data type",              "Cross-section (film-level)"),
        ("Observations (N)",       "4,012"),
        ("Preferred model column", "Table 2, Model 2 (m.pooled.sd)"),
        ("Dependent variable",     "sd_pooled"),
        ("Focal IV",               "FLead (binary, 1 = female lead actor)"),
        ("Controls",               "mean_pooled, kim_violence_gore, kim_sex_nudity, kim_language, major, log_bom_opening_theaters, genres_count"),
        ("Fixed effects",          "bom_year (27 levels) + bom_open_month (12 levels) + 22 genre dummies (treatment coding, drop_first=True)"),
        ("Clustered SEs",          "None (standard OLS SEs — confirmed from R source)"),
        ("Weights",                "None"),
        ("Baseline coefficient",   "0.0468 (FLead)"),
        ("Baseline SE",            "0.0080"),
        ("Baseline p-value",       "<0.001 (p = 6.23e-09)"),
        ("Published value",        "0.047** (Table 2, Model 2; rounds to 0.0468)"),
        ("R-squared",              "0.5620"),
        ("Key variables (4)",      "kim_violence_gore, kim_sex_nudity, kim_language, log_bom_opening_theaters"),
        ("MAR control",            "genres_count"),
        ("Mechanisms",             "MCAR, MAR, NMAR"),
        ("Missingness proportions","1%, 5%, 10%, 20%, 30%, 40%, 50%"),
        ("Methods",                "LD, Mean, Reg, Iter, RF, DL, MILGBM"),
        ("Iterations per scenario","30"),
        ("MAR/NMAR strength",      "1.5"),
        ("M (multiple imputation)","5"),
        ("DL method",              "TensorFlow 2.20.0, MLP 32->Dropout(0.1)->16->1, ReLU, Adam lr=0.005, EarlyStopping"),
        ("MILGBM method",          "LightGBM 4.6.0, MICE M=5, Rubin's Rules, 30 estimators, max_depth=4"),
        ("Full run start",         "2026-03-31 11:00:51"),
        ("Full run end",           "2026-04-01 00:56:51"),
        ("Total runs",             "17,640 (0 errors)"),
        ("QC status",              "PASSED (2026-04-01)"),
    ],
}


# ---------------------------------------------------------------------------
# Task 3 — Paper_Info_Record.pdf
# ---------------------------------------------------------------------------
def _ascii(text: str) -> str:
    """Replace non-latin-1 characters with ASCII equivalents for fpdf2 core fonts."""
    return (text
            .replace("\u2014", "--").replace("\u2013", "-")
            .replace("\u2018", "'").replace("\u2019", "'")
            .replace("\u201c", '"').replace("\u201d", '"')
            .replace("\u00d7", "x").replace("\u2248", "~")
            .encode("latin-1", errors="replace").decode("latin-1"))


def generate_pdf(paper_id: str, paper_dir: Path, pir_txt: Path):
    pdf = FPDF()
    pdf.set_margins(20, 20, 20)
    pdf.add_page()

    # Title block
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(30, 80, 160)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, f"Paper Information Record - Paper {paper_id}", new_x="LMARGIN", new_y="NEXT", fill=True, align="C")
    pdf.ln(4)

    # Fields from structured data
    pdf.set_text_color(0, 0, 0)
    for field, value in PIR_DATA[paper_id]:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(240, 240, 248)
        pdf.cell(65, 7, _ascii(field), border=1, fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(255, 255, 255)
        pdf.multi_cell(0, 7, _ascii(value), border=1)
        pdf.set_x(pdf.l_margin)

    # Append full PIR text as appendix
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Appendix: Full Paper_Info_Record.txt", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 7)
    pdf.ln(2)
    text = pir_txt.read_text(encoding="utf-8")
    for line in text.splitlines():
        pdf.cell(0, 4, _ascii(line[:120]), new_x="LMARGIN", new_y="NEXT")

    out = paper_dir / "Paper_Info_Record.pdf"
    pdf.output(str(out))
    print(f"  PDF written: {out}")
    return out


# ---------------------------------------------------------------------------
# Task 5 — paper_info.xlsx
# ---------------------------------------------------------------------------
def generate_paper_info_xlsx(paper_id: str, paper_dir: Path) -> Path:
    rows = PIR_DATA[paper_id]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Paper_{paper_id}_Info"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E50A0")
    label_fill = PatternFill("solid", fgColor="F0F0F8")
    label_font = Font(bold=True)

    ws.append(["Field", "Value"])
    ws["A1"].font = header_font; ws["A1"].fill = header_fill
    ws["B1"].font = header_font; ws["B1"].fill = header_fill

    for field, value in rows:
        ws.append([field, value])
        row = ws.max_row
        ws[f"A{row}"].font = label_font
        ws[f"A{row}"].fill = label_fill
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    out = paper_dir / "full_run" / f"paper_info_{paper_id}.xlsx"
    wb.save(str(out))
    print(f"  paper_info written: {out}")
    return out


# ---------------------------------------------------------------------------
# Task 5 — regression_results.xlsx
# ---------------------------------------------------------------------------
def generate_regression_results_xlsx(paper_id: str, paper_dir: Path, workbook: Path) -> Path:
    print(f"  Reading workbook: {workbook}")
    # Baseline info from PIR_DATA
    pir = dict(PIR_DATA[paper_id])

    wb_out = openpyxl.Workbook()

    # Sheet 1: Baseline
    ws_base = wb_out.active
    ws_base.title = "Baseline_Regression"
    base_rows = [
        ["Field", "Value"],
        ["Paper ID", paper_id],
        ["Dependent variable", pir["Dependent variable"]],
        ["Focal IV", pir["Focal IV"]],
        ["Controls", pir["Controls"]],
        ["Fixed effects", pir["Fixed effects"]],
        ["Clustered SEs", pir["Clustered SEs"]],
        ["Observations (N)", pir["Observations (N)"]],
        ["Baseline coefficient", pir["Baseline coefficient"]],
        ["Baseline SE", pir["Baseline SE"]],
        ["Baseline p-value", pir["Baseline p-value"]],
        ["Published value", pir["Published value"]],
        ["R-squared", pir["R-squared"]],
    ]
    for row in base_rows:
        ws_base.append(row)
    ws_base["A1"].font = Font(bold=True); ws_base["B1"].font = Font(bold=True)
    ws_base.column_dimensions["A"].width = 28
    ws_base.column_dimensions["B"].width = 80

    # Sheet 2: Coef_Stability_Summary from the main workbook
    try:
        df_css = pd.read_excel(workbook, sheet_name="Coef_Stability_Summary")
        ws_css = wb_out.create_sheet("Coef_Stability_Summary")
        # Write header
        ws_css.append(list(df_css.columns))
        for cell in ws_css[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1E50A0")
            cell.font = Font(bold=True, color="FFFFFF")
        for _, row in df_css.iterrows():
            ws_css.append([str(v) if pd.isna(v) is False else "" for v in row])
        print(f"  Coef_Stability_Summary: {len(df_css)} rows written")
    except Exception as e:
        print(f"  WARNING: Could not read Coef_Stability_Summary: {e}")

    # Sheet 3: B_prop pivot — focal coefficient by method x proportion for first key var
    try:
        df = pd.read_excel(workbook, sheet_name="Coef_Stability_Summary")
        focal_var = df["KeyVar"].iloc[0]
        df_pivot = df[df["KeyVar"] == focal_var].pivot_table(
            index="Method", columns="Proportion", values="B_prop", aggfunc="mean"
        )
        ws_piv = wb_out.create_sheet(f"B_prop_{focal_var}")
        ws_piv.append(["Method \\ Proportion"] + list(df_pivot.columns))
        for cell in ws_piv[1]:
            cell.font = Font(bold=True)
        for method, row in df_pivot.iterrows():
            ws_piv.append([method] + [round(v, 1) if pd.notna(v) else "" for v in row])
        print(f"  B_prop pivot for {focal_var} written")
    except Exception as e:
        print(f"  WARNING: Could not generate B_prop pivot: {e}")

    out = paper_dir / "full_run" / f"regression_results_{paper_id}.xlsx"
    wb_out.save(str(out))
    print(f"  regression_results written: {out}")
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    root_exports = []

    for paper_id, cfg in PAPERS.items():
        paper_dir = cfg["paper_dir"]
        workbook  = cfg["workbook"]
        pir_txt   = cfg["pir_txt"]

        print(f"\n=== Paper {paper_id} ===")

        # PDF
        pdf_path = generate_pdf(paper_id, paper_dir, pir_txt)

        # paper_info.xlsx
        pi_path = generate_paper_info_xlsx(paper_id, paper_dir)

        # regression_results.xlsx
        rr_path = generate_regression_results_xlsx(paper_id, paper_dir, workbook)

        root_exports.extend([pdf_path, pi_path, rr_path])

    # Copy flat exports to repo root (paper_analysis_output level)
    import shutil
    root_out = WEEK2.parent  # Oxford_MonteCarlo root
    for src in root_exports:
        dst = root_out / src.name
        shutil.copy2(src, dst)
        print(f"  Root copy: {dst.name}")

    print("\nDone.")


if __name__ == "__main__":
    main()
